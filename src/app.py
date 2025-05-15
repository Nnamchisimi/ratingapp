from flask import Flask, request, jsonify
from flask_cors import CORS
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime
import webbrowser

app = Flask(__name__)
CORS(app)

EXCEL_FILE = "feedback_weights.xlsx"

# Create Excel file if not exists
def initialize_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Feedback"
        ws.append(["Timestamp", "Emoji", "Weight", "Suggestion"])
        wb.save(EXCEL_FILE)

# Home route (prevents 404 at root)
@app.route("/")
def home():
    return "Feedback API is running."

# Save feedback
@app.route("/save-feedback", methods=["POST"])
def save_feedback():
    initialize_excel()
    data = request.json
    emoji = data.get("emoji")
    weight = data.get("weight")
    suggestion = data.get("suggestion", "")

    if not emoji or weight is None:
        return jsonify({"error": "Missing emoji or weight"}), 400

    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb["Feedback"]
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws.append([timestamp, emoji, weight, suggestion])
        wb.save(EXCEL_FILE)
        return jsonify({"message": "Feedback saved"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

if __name__ == '__main__':
    webbrowser.open("http://127.0.0.1:4000")
    app.run(port=4000)
